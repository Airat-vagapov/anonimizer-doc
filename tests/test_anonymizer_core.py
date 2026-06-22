from __future__ import annotations

import io
from pathlib import Path

import openpyxl
import pytest

from anonymizer_core import (
    Anonymizer,
    anonymize_workbook,
    build_decrypted_output_path,
    build_output_paths,
    deanonymize_text,
    deanonymize_workbook,
    infer_mapping_path,
    iter_existing_cells,
    load_reverse_mapping,
    save_mapping_workbook,
    validate_excel_path,
)


# ---------------------------------------------------------------------------
# Anonymizer — unit tests
# ---------------------------------------------------------------------------


class TestAnonymizerDeterminism:
    def test_same_name_gets_same_replacement(self):
        a = Anonymizer()
        first = a.anonymize_text("Иванов Иван Иванович")
        second = a.anonymize_text("Иванов Иван Иванович")
        assert first == second

    def test_same_login_gets_same_replacement(self):
        a = Anonymizer()
        first = a.anonymize_text("ivan.petrov")
        second = a.anonymize_text("ivan.petrov")
        assert first == second

    def test_different_names_get_different_replacements(self):
        a = Anonymizer()
        r1 = a.anonymize_text("Иванов Иван Иванович")
        r2 = a.anonymize_text("Петрова Мария Сергеевна")
        assert r1 != r2

    def test_different_logins_get_different_replacements(self):
        a = Anonymizer()
        r1 = a.anonymize_text("ivan.petrov")
        r2 = a.anonymize_text("anna.sidorova")
        assert r1 != r2

    def test_name_index_does_not_increment_on_repeat(self):
        a = Anonymizer()
        a.anonymize_text("Иванов Иван Иванович")
        a.anonymize_text("Иванов Иван Иванович")
        assert a._name_index == 1

    def test_login_index_does_not_increment_on_repeat(self):
        a = Anonymizer()
        a.anonymize_text("ivan.petrov")
        a.anonymize_text("ivan.petrov")
        assert a._login_index == 1

    def test_login_numbering_has_no_gaps(self):
        a = Anonymizer()
        a.anonymize_text("ivan.petrov")
        a.anonymize_text("ivan.petrov")  # repeat — should not consume a slot
        a.anonymize_text("anna.sidorova")
        assert a.login_map["anna.sidorova"] == "user0002.masked"


class TestAnonymizerTextPatterns:
    def test_fio_three_words_replaced(self):
        a = Anonymizer()
        result = a.anonymize_text("Выполнял Иванов Иван Иванович")
        assert "Иванов Иван Иванович" not in result

    def test_fio_two_words_replaced(self):
        a = Anonymizer()
        result = a.anonymize_text("Автор: Мария Петрова")
        assert "Мария Петрова" not in result

    def test_jira_login_replaced(self):
        a = Anonymizer()
        result = a.anonymize_text("Задача назначена ivan.petrov")
        assert "ivan.petrov" not in result
        assert "masked" in result

    def test_plain_text_unchanged(self):
        a = Anonymizer()
        text = "Задача по проекту завершена"
        assert a.anonymize_text(text) == text

    def test_empty_string_unchanged(self):
        a = Anonymizer()
        assert a.anonymize_text("") == ""

    def test_mixed_content(self):
        a = Anonymizer()
        text = "Ответственный Иванов Иван Иванович (ivan.ivanov) завершил задачу"
        result = a.anonymize_text(text)
        assert "Иванов Иван Иванович" not in result
        assert "ivan.ivanov" not in result
        assert "masked" in result

    def test_original_names_not_in_output(self):
        a = Anonymizer()
        names = ["Смирнов Алексей Игоревич", "Козлова Дарья Андреевна", "Новиков Никита Олегович"]
        for name in names:
            result = a.anonymize_text(name)
            assert name not in result

    def test_fake_login_format(self):
        a = Anonymizer()
        a.anonymize_text("ivan.petrov")
        replacement = a.login_map["ivan.petrov"]
        assert replacement.startswith("user")
        assert replacement.endswith(".masked")


# ---------------------------------------------------------------------------
# deanonymize_text
# ---------------------------------------------------------------------------


class TestDeanonymizeText:
    def test_restores_single_value(self):
        replacements = [("user0001.masked", "ivan.petrov")]
        assert deanonymize_text("user0001.masked выполнил задачу", replacements) == "ivan.petrov выполнил задачу"

    def test_longer_match_takes_priority(self):
        # "Алексей Александров 1" must not be partially replaced by "Алексей Александров"
        replacements = sorted(
            [
                ("Алексей Александров 1", "Иванов Иван Иванович"),
                ("Алексей Александров", "Петров Пётр"),
            ],
            key=lambda x: len(x[0]),
            reverse=True,
        )
        result = deanonymize_text("Алексей Александров 1", replacements)
        assert result == "Иванов Иван Иванович"

    def test_unchanged_when_no_match(self):
        replacements = [("user0001.masked", "ivan.petrov")]
        text = "обычный текст"
        assert deanonymize_text(text, replacements) == text


# ---------------------------------------------------------------------------
# Round-trip: anonymize → deanonymize produces original
# ---------------------------------------------------------------------------


def _make_xlsx_bytes(rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_tmp_xlsx(tmp_path: Path, rows: list[list], name: str = "input.xlsx") -> Path:
    p = tmp_path / name
    p.write_bytes(_make_xlsx_bytes(rows))
    return p


class TestRoundTrip:
    def test_full_roundtrip_restores_original(self, tmp_path):
        rows = [
            ["Сотрудник", "Логин", "Задача"],
            ["Иванов Иван Иванович", "ivan.ivanov", "Разработка"],
            ["Петрова Мария Сергеевна", "maria.petrova", "Тестирование"],
        ]
        input_path = _write_tmp_xlsx(tmp_path, rows)
        anon_path = tmp_path / "anon.xlsx"
        mapping_path = tmp_path / "mapping.xlsx"

        _, anonymizer = anonymize_workbook(input_path, anon_path)
        save_mapping_workbook(mapping_path, anonymizer)

        decrypted_path = tmp_path / "decrypted.xlsx"
        deanonymize_workbook(anon_path, mapping_path, decrypted_path)

        orig_wb = openpyxl.load_workbook(input_path, data_only=True)
        dec_wb = openpyxl.load_workbook(decrypted_path, data_only=True)

        orig_values = [[c.value for c in row] for row in orig_wb.active.iter_rows()]
        dec_values = [[c.value for c in row] for row in dec_wb.active.iter_rows()]
        assert orig_values == dec_values

    def test_anonymized_file_differs_from_original(self, tmp_path):
        rows = [["Иванов Иван Иванович", "ivan.ivanov"]]
        input_path = _write_tmp_xlsx(tmp_path, rows)
        anon_path = tmp_path / "anon.xlsx"

        anonymize_workbook(input_path, anon_path)

        orig_wb = openpyxl.load_workbook(input_path, data_only=True)
        anon_wb = openpyxl.load_workbook(anon_path, data_only=True)
        orig_vals = [c.value for row in orig_wb.active.iter_rows() for c in row]
        anon_vals = [c.value for row in anon_wb.active.iter_rows() for c in row]
        assert orig_vals != anon_vals

    def test_empty_cells_survive_roundtrip(self, tmp_path):
        rows = [["Иванов Иван Иванович", None, ""], [None, "ivan.ivanov", None]]
        input_path = _write_tmp_xlsx(tmp_path, rows)
        anon_path = tmp_path / "anon.xlsx"
        mapping_path = tmp_path / "mapping.xlsx"

        _, anonymizer = anonymize_workbook(input_path, anon_path)
        save_mapping_workbook(mapping_path, anonymizer)

        decrypted_path = tmp_path / "decrypted.xlsx"
        deanonymize_workbook(anon_path, mapping_path, decrypted_path)

        dec_wb = openpyxl.load_workbook(decrypted_path, data_only=True)
        dec_ws = dec_wb.active
        assert dec_ws.cell(1, 1).value == "Иванов Иван Иванович"

    def test_repeated_name_maps_to_single_replacement(self, tmp_path):
        rows = [
            ["Иванов Иван Иванович"],
            ["Иванов Иван Иванович"],
            ["Иванов Иван Иванович"],
        ]
        input_path = _write_tmp_xlsx(tmp_path, rows)
        anon_path = tmp_path / "anon.xlsx"

        _, anonymizer = anonymize_workbook(input_path, anon_path)
        assert len(anonymizer.full_name_map) == 1

        anon_wb = openpyxl.load_workbook(anon_path, data_only=True)
        values = [row[0].value for row in anon_wb.active.iter_rows()]
        assert values[0] == values[1] == values[2]


# ---------------------------------------------------------------------------
# Path helpers
# ---------------------------------------------------------------------------


class TestPathHelpers:
    def test_build_output_paths(self, tmp_path):
        p = tmp_path / "report.xlsx"
        anon, mapping = build_output_paths(p)
        assert anon.name == "report_anonymized.xlsx"
        assert mapping.name == "report_mapping.xlsx"
        assert anon.parent == mapping.parent

    def test_build_decrypted_output_path(self, tmp_path):
        p = tmp_path / "report_anonymized" / "report_anonymized.xlsx"
        out = build_decrypted_output_path(p)
        assert out.name == "report_anonymized_decrypted.xlsx"

    def test_infer_mapping_path_found(self, tmp_path):
        anon_dir = tmp_path / "report_anonymized"
        anon_dir.mkdir()
        mapping = anon_dir / "report_mapping.xlsx"
        mapping.touch()
        anon_file = anon_dir / "report_anonymized.xlsx"
        result = infer_mapping_path(anon_file)
        assert result == mapping

    def test_infer_mapping_path_not_found(self, tmp_path):
        anon_file = tmp_path / "report_anonymized.xlsx"
        with pytest.raises(FileNotFoundError):
            infer_mapping_path(anon_file)

    def test_validate_excel_path_wrong_extension(self, tmp_path):
        p = tmp_path / "file.csv"
        p.touch()
        with pytest.raises(ValueError):
            validate_excel_path(p)

    def test_validate_excel_path_not_found(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            validate_excel_path(tmp_path / "missing.xlsx")

    def test_validate_excel_path_valid(self, tmp_path):
        p = tmp_path / "file.xlsx"
        p.write_bytes(_make_xlsx_bytes([["test"]]))
        result = validate_excel_path(p)
        assert result == p.resolve()


# ---------------------------------------------------------------------------
# iter_existing_cells
# ---------------------------------------------------------------------------


class TestIterExistingCells:
    def test_skips_empty_cells(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "hello"
        ws["C3"] = "world"
        cells = list(iter_existing_cells(ws))
        values = {c.value for c in cells if c.value is not None}
        assert values == {"hello", "world"}
