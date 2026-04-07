from __future__ import annotations

import argparse
import re
import subprocess
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook

FIRST_NAMES = [
    "Алексей",
    "Андрей",
    "Анна",
    "Виктор",
    "Дарья",
    "Екатерина",
    "Ирина",
    "Мария",
    "Никита",
    "Ольга",
    "Сергей",
    "Татьяна",
]

LAST_NAMES = [
    "Александров",
    "Белов",
    "Волков",
    "Громов",
    "Егоров",
    "Захаров",
    "Ильин",
    "Крылов",
    "Морозов",
    "Назаров",
    "Орлов",
    "Соколов",
]

MIDDLE_NAMES = [
    "Алексеевич",
    "Андреевич",
    "Викторович",
    "Дмитриевич",
    "Игоревич",
    "Олегович",
    "Сергеевич",
    "Алексеевна",
    "Андреевна",
    "Викторовна",
    "Дмитриевна",
    "Игоревна",
    "Олеговна",
    "Сергеевна",
]

FIO_PATTERN = re.compile(
    r"\b[А-ЯЁ][а-яё]+(?:-[А-ЯЁ][а-яё]+)?(?:\s+[А-ЯЁ][а-яё]+(?:-[А-ЯЁ][а-яё]+)?){1,2}\b"
)
JIRA_LOGIN_PATTERN = re.compile(r"\b[a-z][a-z0-9_-]*\.[a-z][a-z0-9_-]*\b", re.IGNORECASE)
SUPPORTED_EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}
MAPPING_SHEET_NAME = "Расшифровка"


class Anonymizer:
    def __init__(self) -> None:
        self.full_name_map: dict[str, str] = {}
        self.login_map: dict[str, str] = {}
        self._used_names: set[str] = set()
        self._name_index = 0
        self._login_index = 0

    def anonymize_text(self, value: str) -> str:
        value = self._replace_full_names(value)
        value = self._replace_logins(value)
        return value

    def _replace_full_names(self, value: str) -> str:
        return FIO_PATTERN.sub(self._full_name_replacer, value)

    def _replace_logins(self, value: str) -> str:
        return JIRA_LOGIN_PATTERN.sub(self._login_replacer, value)

    def _full_name_replacer(self, match: re.Match[str]) -> str:
        original_name = match.group(0)
        return self.full_name_map.setdefault(original_name, self._generate_fake_name(original_name))

    def _login_replacer(self, match: re.Match[str]) -> str:
        original_login = match.group(0)
        return self.login_map.setdefault(original_login, self._generate_fake_login())

    def _generate_fake_name(self, original_name: str) -> str:
        parts_count = len(original_name.split())

        while True:
            first_name = FIRST_NAMES[self._name_index % len(FIRST_NAMES)]
            last_name = LAST_NAMES[(self._name_index // len(FIRST_NAMES)) % len(LAST_NAMES)]
            middle_name = MIDDLE_NAMES[
                (self._name_index // (len(FIRST_NAMES) * len(LAST_NAMES))) % len(MIDDLE_NAMES)
            ]
            self._name_index += 1

            if parts_count == 2:
                candidate = f"{first_name} {last_name}"
            else:
                candidate = f"{last_name} {first_name} {middle_name}"

            if candidate not in self._used_names:
                self._used_names.add(candidate)
                return candidate

    def _generate_fake_login(self) -> str:
        self._login_index += 1
        return f"user{self._login_index:04d}.masked"


def remove_hyperlink(cell) -> None:
    if cell.hyperlink is not None:
        cell.hyperlink = None


def validate_excel_path(input_path: Path) -> Path:
    resolved_path = input_path.expanduser().resolve()
    if resolved_path.suffix.lower() not in SUPPORTED_EXCEL_SUFFIXES:
        raise ValueError("Поддерживаются только Excel-файлы: .xlsx, .xlsm, .xltx, .xltm")
    if not resolved_path.exists():
        raise FileNotFoundError(f"Файл не найден: {resolved_path}")
    return resolved_path


def anonymize_workbook(input_path: Path, output_path: Path) -> tuple[Path, Anonymizer]:
    workbook = load_workbook(input_path)
    anonymizer = Anonymizer()

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                remove_hyperlink(cell)
                if isinstance(cell.value, str):
                    cell.value = anonymizer.anonymize_text(cell.value)

    workbook.save(output_path)
    return output_path, anonymizer


def save_mapping_workbook(
    output_path: Path, anonymizer: Anonymizer, sheet_name: str = MAPPING_SHEET_NAME
) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(["Тип", "Оригинал", "Замена"])

    for original_name, fake_name in anonymizer.full_name_map.items():
        worksheet.append(["ФИО", original_name, fake_name])

    for original_login, fake_login in anonymizer.login_map.items():
        worksheet.append(["Логин", original_login, fake_login])

    for column in ("A", "B", "C"):
        worksheet.column_dimensions[column].width = 32

    workbook.save(output_path)
    return output_path


def load_reverse_mapping(mapping_path: Path) -> list[tuple[str, str]]:
    workbook = load_workbook(mapping_path, read_only=True, data_only=True)
    worksheet = workbook[MAPPING_SHEET_NAME] if MAPPING_SHEET_NAME in workbook.sheetnames else workbook.active
    replacements: list[tuple[str, str]] = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue

        _, original_value, masked_value = row[:3]
        if not isinstance(original_value, str) or not isinstance(masked_value, str):
            continue
        if not masked_value:
            continue
        replacements.append((masked_value, original_value))

    workbook.close()
    return sorted(replacements, key=lambda item: len(item[0]), reverse=True)


def deanonymize_text(value: str, replacements: Iterable[tuple[str, str]]) -> str:
    restored_value = value
    for masked_value, original_value in replacements:
        restored_value = restored_value.replace(masked_value, original_value)
    return restored_value


def deanonymize_workbook(input_path: Path, mapping_path: Path, output_path: Path) -> Path:
    workbook = load_workbook(input_path)
    replacements = load_reverse_mapping(mapping_path)

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    cell.value = deanonymize_text(cell.value, replacements)

    workbook.save(output_path)
    return output_path


def build_output_dir(input_path: Path, suffix: str) -> Path:
    return input_path.parent / f"{input_path.stem}_{suffix}"


def build_output_paths(input_path: Path) -> tuple[Path, Path]:
    output_dir = build_output_dir(input_path, "anonymized")
    anonymized_path = output_dir / f"{input_path.stem}_anonymized{input_path.suffix}"
    mapping_path = output_dir / f"{input_path.stem}_mapping.xlsx"
    return anonymized_path, mapping_path


def build_decrypted_output_path(input_path: Path) -> Path:
    output_dir = build_output_dir(input_path, "decrypted")
    return output_dir / f"{input_path.stem}_decrypted{input_path.suffix}"


def infer_mapping_path(input_path: Path) -> Path:
    stem = input_path.stem
    original_stem = stem.removesuffix("_anonymized")
    candidate = input_path.parent / f"{original_stem}_mapping.xlsx"
    if candidate.exists():
        return candidate
    raise FileNotFoundError(
        "Не удалось автоматически найти файл расшифровки. "
        "Передайте путь через аргумент --mapping-file."
    )


def choose_input_file(prompt: str = "Выберите Excel-файл для обезличивания") -> Path:
    script = """
    set selectedFile to choose file with prompt "%s"
    POSIX path of selectedFile
    """ % prompt
    try:
        result = subprocess.run(
            ["osascript", "-e", script],
            capture_output=True,
            text=True,
            check=True,
        )
        selected_path = result.stdout.strip()
        if selected_path:
            return validate_excel_path(Path(selected_path))
    except (subprocess.CalledProcessError, FileNotFoundError):
        pass

    manual_path = input("Введите путь до Excel-файла: ").strip()
    if not manual_path:
        raise ValueError("Не выбран входной файл")
    return validate_excel_path(Path(manual_path))


def choose_mapping_file() -> Path:
    return choose_input_file("Выберите файл расшифровки mapping.xlsx")


def choose_mode() -> bool:
    script = """
    set selectedAction to button returned of (display dialog "Что нужно сделать?" buttons {"Расшифровать", "Обезличить"} default button "Обезличить")
    selectedAction
    """
    try:
        result = subprocess.run(
            ["osascript", "-e", script],
            capture_output=True,
            text=True,
            check=True,
        )
        return result.stdout.strip() == "Расшифровать"
    except (subprocess.CalledProcessError, FileNotFoundError):
        pass

    try:
        choice = input(
            "Выберите режим: 1 - обезличить файл, 2 - расшифровать файл по mapping.xlsx: "
        ).strip()
    except EOFError:
        choice = ""

    if choice == "2":
        return True
    if choice in {"", "1"}:
        return False
    raise ValueError("Поддерживаются только режимы 1 и 2")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Обезличивание и расшифровка Excel-файла")
    parser.add_argument("input_file", nargs="?", help="Путь до Excel-файла")
    parser.add_argument(
        "--decrypt",
        action="store_true",
        help="Расшифровать ранее обезличенный Excel-файл",
    )
    parser.add_argument(
        "--mapping-file",
        help="Путь до файла соответствий *_mapping.xlsx для режима --decrypt",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    decrypt_mode = args.decrypt

    if not args.input_file and not args.decrypt:
        decrypt_mode = choose_mode()

    input_path = (
        validate_excel_path(Path(args.input_file))
        if args.input_file
        else choose_input_file(
            "Выберите Excel-файл для расшифровки"
            if decrypt_mode
            else "Выберите Excel-файл для обезличивания"
        )
    )

    if decrypt_mode:
        if args.mapping_file:
            mapping_path = validate_excel_path(Path(args.mapping_file))
        else:
            try:
                mapping_path = infer_mapping_path(input_path)
            except FileNotFoundError:
                mapping_path = choose_mapping_file()
        output_path = build_decrypted_output_path(input_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        decrypted_file = deanonymize_workbook(input_path, mapping_path, output_path)
        print(f"Расшифрованный файл сохранен: {decrypted_file}")
        print(f"Использован файл расшифровки: {mapping_path}")
        return 0

    output_path, mapping_path = build_output_paths(input_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    anonymized_file, anonymizer = anonymize_workbook(input_path, output_path)
    mapping_file = save_mapping_workbook(mapping_path, anonymizer)
    print(f"Обезличенный файл сохранен: {anonymized_file}")
    print(f"Файл расшифровки сохранен: {mapping_file}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
