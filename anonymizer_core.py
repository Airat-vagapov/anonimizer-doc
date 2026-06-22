from __future__ import annotations

import re
from pathlib import Path
from typing import Callable, Iterable

from openpyxl import Workbook, load_workbook

FIRST_NAMES = [
    "Алексей", "Александр", "Алина", "Андрей", "Анна", "Антон",
    "Борис", "Валентина", "Валерий", "Виктор", "Виктория", "Владимир",
    "Дарья", "Денис", "Диана", "Дмитрий", "Екатерина", "Елена",
    "Иван", "Игорь", "Ирина", "Кирилл", "Ксения", "Константин",
    "Лариса", "Людмила", "Максим", "Мария", "Михаил", "Надежда",
    "Никита", "Николай", "Оксана", "Олег", "Ольга", "Павел",
    "Полина", "Роман", "Светлана", "Сергей", "Тамара", "Татьяна",
    "Юлия", "Юрий", "Яна",
]

LAST_NAMES = [
    "Александров", "Алексеев", "Афанасьев", "Баранов", "Белов", "Борисов",
    "Васильев", "Виноградов", "Волков", "Воробьёв", "Герасимов", "Григорьев",
    "Громов", "Данилов", "Егоров", "Ефимов", "Захаров", "Зайцев",
    "Иванов", "Ильин", "Козлов", "Комаров", "Кузнецов", "Кузьмин",
    "Крылов", "Лазарев", "Лебедев", "Макаров", "Медведев", "Михайлов",
    "Морозов", "Назаров", "Никитин", "Николаев", "Новиков", "Орлов",
    "Павлов", "Петров", "Попов", "Романов", "Семёнов", "Сидоров",
    "Смирнов", "Соколов", "Степанов", "Тихонов", "Федоров", "Фролов",
    "Харитонов", "Чистяков", "Щербаков",
]

MIDDLE_NAMES = [
    "Алексеевич", "Алексеевна",
    "Александрович", "Александровна",
    "Андреевич", "Андреевна",
    "Борисович", "Борисовна",
    "Валерьевич", "Валерьевна",
    "Викторович", "Викторовна",
    "Владимирович", "Владимировна",
    "Дмитриевич", "Дмитриевна",
    "Евгеньевич", "Евгеньевна",
    "Игоревич", "Игоревна",
    "Константинович", "Константиновна",
    "Михайлович", "Михайловна",
    "Николаевич", "Николаевна",
    "Олегович", "Олеговна",
    "Павлович", "Павловна",
    "Петрович", "Петровна",
    "Сергеевич", "Сергеевна",
    "Юрьевич", "Юрьевна",
]

# For 3-word matches the third word must end in a patronymic suffix (-вич/-вна/-ична)
# to avoid replacing arbitrary capitalized Cyrillic phrases like "Отдел Продаж Москва".
_CYRILLIC_WORD = r"[А-ЯЁ][а-яё]+(?:-[А-ЯЁ][а-яё]+)?"
_PATRONYMIC_WORD = r"[А-ЯЁ][а-яё]*(?:вич|вна|ична)"
FIO_PATTERN = re.compile(
    r"\b" + _CYRILLIC_WORD + r"\s+" + _CYRILLIC_WORD + r"(?:\s+" + _PATRONYMIC_WORD + r")?\b"
)
JIRA_LOGIN_PATTERN = re.compile(r"\b[a-z][a-z0-9_-]*\.[a-z][a-z0-9_-]*\b", re.IGNORECASE)
SUPPORTED_EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}
MAPPING_SHEET_NAME = "Расшифровка"
MAPPING_COLUMN_WIDTH = 32


class Anonymizer:
    def __init__(self) -> None:
        self.full_name_map: dict[str, str] = {}
        self.login_map: dict[str, str] = {}
        self._name_index = 0
        self._login_index = 0

    def anonymize_text(self, value: str) -> str:
        value = self._replace_full_names(value)
        value = self._replace_logins(value)
        return value

    def _replace_full_names(self, value: str) -> str:
        return FIO_PATTERN.sub(self._full_name_replacer, value)

    def _replace_logins(self, value: str) -> str:
        return JIRA_LOGIN_PATTERN.sub(self._login_replacer, value)

    def _full_name_replacer(self, match: re.Match[str]) -> str:
        original_name = match.group(0)
        if original_name not in self.full_name_map:
            self.full_name_map[original_name] = self._generate_fake_name(original_name)
        return self.full_name_map[original_name]

    def _login_replacer(self, match: re.Match[str]) -> str:
        original_login = match.group(0)
        if original_login not in self.login_map:
            self.login_map[original_login] = self._generate_fake_login()
        return self.login_map[original_login]

    def _generate_fake_name(self, original_name: str) -> str:
        parts_count = len(original_name.split())
        current_index = self._name_index
        self._name_index += 1

        first_name = FIRST_NAMES[current_index % len(FIRST_NAMES)]
        last_name = LAST_NAMES[(current_index // len(FIRST_NAMES)) % len(LAST_NAMES)]

        if parts_count == 2:
            cycle = current_index // (len(FIRST_NAMES) * len(LAST_NAMES))
            candidate = f"{first_name} {last_name}"
            if cycle:
                candidate = f"{candidate} {cycle}"
        else:
            middle_name = MIDDLE_NAMES[
                (current_index // (len(FIRST_NAMES) * len(LAST_NAMES))) % len(MIDDLE_NAMES)
            ]
            cycle = current_index // (len(FIRST_NAMES) * len(LAST_NAMES) * len(MIDDLE_NAMES))
            candidate = f"{last_name} {first_name} {middle_name}"
            if cycle:
                candidate = f"{candidate} {cycle}"

        return candidate

    def _generate_fake_login(self) -> str:
        self._login_index += 1
        return f"user{self._login_index:04d}.masked"


def remove_hyperlink(cell) -> None:
    if cell.hyperlink is not None:
        cell.hyperlink = None


def iter_existing_cells(worksheet):
    # worksheet.iter_rows() allocates a full grid even for sparse sheets;
    # accessing the internal cell store directly is much faster for large workbooks.
    # We fall back to the public API if the internal attribute is unavailable
    # (e.g. after an openpyxl refactor) so the code degrades gracefully.
    cells = getattr(worksheet, "_cells", None)
    if isinstance(cells, dict) and cells:
        return cells.values()

    return (
        cell
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value is not None or cell.hyperlink is not None
    )


def validate_excel_path(input_path: Path) -> Path:
    resolved_path = input_path.expanduser().resolve()
    if resolved_path.suffix.lower() not in SUPPORTED_EXCEL_SUFFIXES:
        raise ValueError("Поддерживаются только Excel-файлы: .xlsx, .xlsm, .xltx, .xltm")
    if not resolved_path.exists():
        raise FileNotFoundError(f"Файл не найден: {resolved_path}")
    return resolved_path


def anonymize_workbook(
    input_path: Path,
    output_path: Path,
    progress: bool = False,
) -> tuple[Path, Anonymizer]:
    workbook = load_workbook(input_path)
    anonymizer = Anonymizer()
    sheets = workbook.worksheets
    total = len(sheets)

    for idx, worksheet in enumerate(sheets, start=1):
        if progress:
            print(f"  Лист {idx}/{total}: {worksheet.title}")
        for cell in iter_existing_cells(worksheet):
            remove_hyperlink(cell)
            if isinstance(cell.value, str):
                cell.value = anonymizer.anonymize_text(cell.value)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path, anonymizer


def save_mapping_workbook(
    output_path: Path, anonymizer: Anonymizer, sheet_name: str = MAPPING_SHEET_NAME
) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(["Тип", "Оригинал", "Замена"])

    for original_name, fake_name in anonymizer.full_name_map.items():
        worksheet.append(["ФИО", original_name, fake_name])

    for original_login, fake_login in anonymizer.login_map.items():
        worksheet.append(["Логин", original_login, fake_login])

    for column in ("A", "B", "C"):
        worksheet.column_dimensions[column].width = MAPPING_COLUMN_WIDTH

    workbook.save(output_path)
    return output_path


def load_reverse_mapping(mapping_path: Path) -> list[tuple[str, str]]:
    workbook = load_workbook(mapping_path, read_only=True, data_only=True)
    worksheet = workbook[MAPPING_SHEET_NAME] if MAPPING_SHEET_NAME in workbook.sheetnames else workbook.active
    replacements: list[tuple[str, str]] = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue

        _, original_value, masked_value = row[:3]
        if not isinstance(original_value, str) or not isinstance(masked_value, str):
            continue
        if not masked_value:
            continue
        replacements.append((masked_value, original_value))

    workbook.close()
    return sorted(replacements, key=lambda item: len(item[0]), reverse=True)


def deanonymize_text(value: str, replacements: Iterable[tuple[str, str]]) -> str:
    restored_value = value
    for masked_value, original_value in replacements:
        restored_value = restored_value.replace(masked_value, original_value)
    return restored_value


def _build_reverse_replacer(replacements: list[tuple[str, str]]) -> Callable[[str], str] | None:
    """Compile one regex that swaps all masked values back in a single pass."""
    if not replacements:
        return None
    # replacements is already sorted longest-first by load_reverse_mapping,
    # so longer tokens win over their substrings without extra logic.
    pattern = re.compile("|".join(re.escape(masked) for masked, _ in replacements))
    lookup = {masked: original for masked, original in replacements}
    return lambda value: pattern.sub(lambda m: lookup[m.group(0)], value)


def deanonymize_workbook(
    input_path: Path,
    mapping_path: Path,
    output_path: Path,
    progress: bool = False,
) -> Path:
    workbook = load_workbook(input_path)
    replacements = load_reverse_mapping(mapping_path)
    apply = _build_reverse_replacer(replacements)
    sheets = workbook.worksheets
    total = len(sheets)

    for idx, worksheet in enumerate(sheets, start=1):
        if progress:
            print(f"  Лист {idx}/{total}: {worksheet.title}")
        if apply is None:
            continue
        for cell in iter_existing_cells(worksheet):
            if isinstance(cell.value, str):
                cell.value = apply(cell.value)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def build_output_dir(input_path: Path, suffix: str) -> Path:
    return input_path.parent / f"{input_path.stem}_{suffix}"


def build_output_paths(input_path: Path) -> tuple[Path, Path]:
    output_dir = build_output_dir(input_path, "anonymized")
    anonymized_path = output_dir / f"{input_path.stem}_anonymized{input_path.suffix}"
    mapping_path = output_dir / f"{input_path.stem}_mapping.xlsx"
    return anonymized_path, mapping_path


def build_decrypted_output_path(input_path: Path) -> Path:
    output_dir = build_output_dir(input_path, "decrypted")
    return output_dir / f"{input_path.stem}_decrypted{input_path.suffix}"


def infer_mapping_path(input_path: Path) -> Path:
    stem = input_path.stem
    original_stem = stem.removesuffix("_anonymized")
    candidate = input_path.parent / f"{original_stem}_mapping.xlsx"
    if candidate.exists():
        return candidate
    raise FileNotFoundError(
        "Не удалось автоматически найти файл расшифровки. "
        "Передайте путь через аргумент --mapping-file."
    )
